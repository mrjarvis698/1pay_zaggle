import os
from os import path
import shutil
import json
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import time

# Open xlsx file
open_sheet = path.exists("cache/opened_sheet.json")
if open_sheet == True :
  opened_sheet_file_path = "cache/opened_sheet.json"
  json_file = open(opened_sheet_file_path)
  data = json.load(json_file)
  xlsx_sheet_check = path.exists(data ['xlsx_file_path'])
  if xlsx_sheet_check == True :
    xlsx_file_path = data ['xlsx_file_path']
  else :
    shutil.rmtree('cache', ignore_errors=True)
    xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
    cache_path = os.path.join(str(os.getcwd()), "cache")
    dictionary = {"xlsx_file_path" : xlsx_file_path}
    json_object = json.dumps(dictionary, indent = 1)
    with open("cache/opened_sheet.json", "w") as outfile:
      outfile.write(json_object)
else :
  xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
  cache_path = os.path.join(str(os.getcwd()), "cache")
  os.mkdir(cache_path)
  dictionary = {"xlsx_file_path" : xlsx_file_path}
  json_object = json.dumps(dictionary, indent = 1)
  with open("cache/opened_sheet.json", "w") as outfile:
    outfile.write(json_object)

# read imported xlsx file path using pandas
input_workbook = pd.read_excel(xlsx_file_path, sheet_name = 'Sheet1', dtype=str)
total_input_rows, total_input_cols = input_workbook.shape
print('Total Cards = ',total_input_rows)

'''
pesudo code for dynamic loading of variables
for heading in input_col:
    input_variables = input_workbook[heading].values.tolist()
    dictionary = {heading : input_variables}
    json_object = json.dumps(dictionary, indent = 1)
    with open("cache/"+ heading +".json", "w") as outfile:
      outfile.write(json_object)
    input_variables = "hemlo" + heading
    print (input_variables)
    input_variables = input_workbook[heading].values.tolist()
    print (input_variables)
'''

input_col = list(input_workbook.columns.values.tolist())

input_xlsx_col_A = input_workbook[input_col[0]].values.tolist()
input_xlsx_col_B = input_workbook[input_col[1]].values.tolist()
input_xlsx_col_C = input_workbook[input_col[2]].values.tolist()
input_xlsx_col_D = input_workbook[input_col[3]].values.tolist()
input_xlsx_col_E = input_workbook[input_col[4]].values.tolist()
input_xlsx_col_F = input_workbook[input_col[5]].values.tolist()
input_xlsx_col_G = input_workbook[input_col[6]].values.tolist()
input_xlsx_col_H = input_workbook[input_col[7]].values.tolist()
input_xlsx_col_I = input_workbook[input_col[8]].values.tolist()
input_xlsx_col_J = input_workbook[input_col[9]].values.tolist()

# get-output sheet to append output
output_sheet = path.exists("Output.xlsx")
if output_sheet == True :
  output_sheet_file_path = "Output.xlsx"
else :
  input_col.append('No.of Transactions')
  input_col.append('Status')
  input_col.append('PG Transaction Number')
  output_headers = input_col
  overall_output = Workbook()
  page = overall_output.active
  page.append(output_headers)
  overall_output.save(filename = 'Output.xlsx')
  output_sheet_file_path = "Output.xlsx"

def output_save():
  global output_wb, entry_list
  entry_list = [[input_xlsx_col_A[x], input_xlsx_col_B[x], input_xlsx_col_C[x], input_xlsx_col_D[x], input_xlsx_col_E[x], input_xlsx_col_F[x], input_xlsx_col_G[x], input_xlsx_col_H[x], input_xlsx_col_I[x], input_xlsx_col_J[x], z + 1, transaction_status, pg_transaction_reference_number]]
  output_wb = load_workbook(output_sheet_file_path)
  page = output_wb.active
  for info in entry_list:
      page.append(info)
  output_wb.save(filename='Output.xlsx')

#for x in range(0, total_input_rows):
    #output_save()

def cal():
  global output_cc_number
  global done_transactions_wb
  global h
  
  output_load_wb = pd.read_excel(output_sheet_file_path, sheet_name = 'Sheet', dtype=str)
  output_col = list(output_load_wb.columns.values.tolist())
  output_cc_number = output_load_wb[output_col[4]].values.tolist()
  done_transactions_wb = output_load_wb[output_col[10]].values.tolist()
  total_output_rows, total_output_cols = output_load_wb.shape
  h = total_output_rows - 1
  print (output_cc_number[h],done_transactions_wb[h])

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")
caps = DesiredCapabilities().CHROME
#caps["pageLoadStrategy"] = "none"
#caps["pageLoadStrategy"] = "eager"
caps["pageLoadStrategy"] = "normal"
driver=webdriver.Chrome(chrome_options=chrome_options, desired_capabilities=caps, executable_path="chromedriver.exe")
driver.maximize_window()

def cc_expiry():
  global expiry_month
  global expiry_year
  global expiry_year1
  global expiry_year2
  global expiry_year3
  global expiry_year4
  workbook_expiry_month = input_xlsx_col_H[x]
  workbook_expiry_year = input_xlsx_col_H[x]
  expiry_month = workbook_expiry_month[:2]
  expiry_year = workbook_expiry_year[5:]
  expiry_year1 = workbook_expiry_year[3]
  expiry_year2 = workbook_expiry_year[4]
  expiry_year3 = workbook_expiry_year[5]
  expiry_year4 = workbook_expiry_year[6]

def start_link():
  driver.get("https://pay.1paypg.in/onepayVAS/earn-more156")

def main():
  global transaction_status, pg_transaction_reference_number
  driver.find_element_by_id("custMobile").click()
  driver.find_element_by_id("custMobile").clear()
  driver.find_element_by_id("custMobile").send_keys(input_xlsx_col_C[x])
  driver.find_element_by_id("udf1").click()
  driver.find_element_by_id("udf1").clear()
  driver.find_element_by_id("udf1").send_keys(input_xlsx_col_A[x] + input_xlsx_col_B[x])
  driver.find_element_by_id("custMail").click()
  driver.find_element_by_id("custMail").clear()
  driver.find_element_by_id("custMail").send_keys("test@test.com")
  driver.find_element_by_id("subMe").click()
  time.sleep(2)
  driver.find_element_by_xpath('//*[@id="tab-menu"]/li/label').click()
  driver.find_element_by_xpath("//div[@id='credit-card-info']/div[4]/div/label").click()
  driver.find_element_by_id("card-number").click()
  driver.find_element_by_id("card-number").send_keys(input_xlsx_col_E[x])
  driver.find_element_by_id("card-holder").click()
  driver.find_element_by_id("card-holder").clear()
  driver.find_element_by_id("card-holder").send_keys(input_xlsx_col_A[x] + input_xlsx_col_B[x])
  driver.find_element_by_id("card-month").click()
  driver.find_element_by_id("card-month").clear()
  driver.find_element_by_id("card-month").send_keys("04")
  driver.find_element_by_id("card-year").click()
  driver.find_element_by_id("card-year").clear()
  driver.find_element_by_id("card-year").send_keys("2023")
  driver.find_element_by_id("card-cvc").click()
  driver.find_element_by_id("card-cvc").clear()
  driver.find_element_by_id("card-cvc").send_keys(input_xlsx_col_G[x])
  driver.find_element_by_id("payNowCC1").click()
  time.sleep(2)
  driver.find_element_by_xpath('//*[@id="tab-B-label"]').click()
  driver.find_element_by_id("expDate").click()
  driver.find_element_by_id("expDate").clear()
  driver.find_element_by_id("expDate").send_keys(expiry_month)
  driver.find_element_by_id("expDate").send_keys(expiry_year1)
  driver.find_element_by_id("expDate").send_keys(expiry_year2)
  driver.find_element_by_id("expDate").send_keys(expiry_year3)
  driver.find_element_by_id("expDate").send_keys(expiry_year4)
  driver.find_element_by_id("pin").click()
  driver.find_element_by_id("pin").clear()
  driver.find_element_by_id("pin").send_keys(input_xlsx_col_F[x])
  driver.find_element_by_xpath("//button[@id='submitButtonIdForPin']/b").click()
  time.sleep(2)
  transaction_status = driver.find_element_by_xpath("/html/body/center/h1").text
  pg_transaction_reference_number = driver.find_element_by_xpath("/html/body/table[2]/tbody/tr[5]/td[2]").text

try:
  cal()
except IndexError:
  for x in range (0 , total_input_rows):
    for z in range (0, 5):
      cc_expiry()
      start_link()
      main()
      output_save()
      print (entry_list)
      time.sleep(0.5)
else:
  last_txncard =  input_workbook[input_workbook[input_col[4]] == output_cc_number[h]].index[0]
  for x in range (last_txncard , total_input_rows):
    for z in range (int(done_transactions_wb[h]), 5):
      cc_expiry()
      start_link()
      main()
      output_save()
      print (entry_list)
      time.sleep(0.5)
    done_transactions_wb[h] = 0

driver.quit()
